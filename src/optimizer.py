import pandas as pd
from ortools.linear_solver import pywraplp
from typing import Tuple, Optional, Dict, List, Any
import logging
from pathlib import Path
import io
import json
import sqlite3
import db_utils
import openpyxl.styles
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Database Tables
TABLE_TEST_DATA = 'test_data'
TABLE_CD_RATES = 'cd_rates'
TABLE_BRANCH_RELATIONSHIPS = 'branch_relationships'
TABLE_ECR_RATES = 'ecr_rates'

# Test Data Columns
class TestDataColumns:
    BRANCH_NAME = 'branch_name'
    ASSOCIATION_NAME = 'association_name'
    HOLDER = 'holder'
    INVESTMENT_TYPE = 'investment_type'
    CURRENT_BALANCE = 'current_balance'

# CD Rates Columns
class CDRatesColumns:
    BANK_NAME = 'bank_name'
    BANK_CODE = 'bank_code'
    CD_TERM = 'cd_term'
    CD_RATE = 'cd_rate'
    CDARS_TERM = 'cdars_term'
    CDARS_RATE = 'cdars_rate'
    SPECIAL = 'special'

# Branch Relationships Columns
class BranchRelationshipsColumns:
    BRANCH_NAME = 'branch_name'
    ALLIANCE_ASSOC_BANK = 'alliance_assoc_bank'
    BANCO_POPULAR = 'banco_popular'
    BANK_UNITED = 'bank_united'
    CITY_NATIONAL = 'city_national'
    ENTERPRISE_BANK_TRUST = 'enterprise_bank_trust'
    FIRST_CITIZENS_BANK = 'first_citizens_bank'
    HARMONY_BANK = 'harmony_bank'
    PACIFIC_PREMIER_BANK = 'pacific_premier_bank'
    PACIFIC_WESTERN = 'pacific_western'
    SOUTHSTATE = 'southstate'
    SUNWEST_BANK = 'sunwest_bank'
    CAPITAL_ONE = 'capital_one'

# ECR Rates Columns
class ECRRatesColumns:
    BANK_NAME = 'bank_name'
    BANK_CODE = 'bank_code'
    ECR_RATE = 'ecr_rate'

def execute_sql_query(query: str, params: Optional[List[Any]] = None) -> pd.DataFrame:
    """
    Execute an SQL query on the database and return results as a DataFrame.
    
    Args:
        query: SQL query string
        params: Optional list of parameters to use in the query
        
    Returns:
        DataFrame containing query results
        
    Raises:
        Exception: If database connection or query execution fails
    """
    try:
        # Connect to the database
        db_path = Path(__file__).parent.parent / "data" / "langston.db"
        conn = sqlite3.connect(db_path)
        
        # Execute query and return results as DataFrame
        if params:
            df = pd.read_sql_query(query, conn, params=params)
        else:
            df = pd.read_sql_query(query, conn)
        conn.close()
        return df
        
    except Exception as e:
        logger.error(f"Error executing SQL query: {str(e)}")
        raise

def get_database_columns() -> Dict[str, List[str]]:
    """
    Connect to the langston.db database and retrieve all column names for each table.
    
    Returns:
        Dict[str, List[str]]: Dictionary mapping table names to their column names
    """
    try:
        # Connect to the database
        db_path = Path(__file__).parent.parent / "data" / "langston.db"
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Get list of tables
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        
        # Dictionary to store table columns
        table_columns = {}
        
        # Get columns for each table
        for table in tables:
            table_name = table[0]
            cursor.execute(f"PRAGMA table_info({table_name});")
            columns = [col[1] for col in cursor.fetchall()]
            table_columns[table_name] = columns
            
        conn.close()
        return table_columns
        
    except Exception as e:
        logger.error(f"Error connecting to database: {str(e)}")
        return {}

# Example SQL queries for common operations
def get_cd_rates(bank_name: Optional[str] = None) -> pd.DataFrame:
    """
    Get CD rates for all banks or a specific bank.
    
    Args:
        bank_name: Optional bank name to filter results
        
    Returns:
        DataFrame with CD rates
    """
    query = f"""
    SELECT {CDRatesColumns.BANK_NAME}, 
           {CDRatesColumns.CD_TERM}, 
           {CDRatesColumns.CD_RATE}
    FROM {TABLE_CD_RATES}
    """
    
    if bank_name:
        query += f" WHERE {CDRatesColumns.BANK_NAME} = '{bank_name}'"
    
    query += f" ORDER BY {CDRatesColumns.BANK_NAME}, {CDRatesColumns.CD_TERM}"
    
    return execute_sql_query(query)

def get_branch_relationships(branch_name: Optional[str] = None) -> pd.DataFrame:
    """
    Get branch relationships for all branches or a specific branch.
    
    Args:
        branch_name: Optional branch name to filter results
        
    Returns:
        DataFrame with branch relationships
    """
    query = f"""
    SELECT *
    FROM {TABLE_BRANCH_RELATIONSHIPS}
    """
    
    params = []
    if branch_name:
        query += f" WHERE {BranchRelationshipsColumns.BRANCH_NAME} = ?"
        params.append(branch_name)
    
    return execute_sql_query(query, params)

def get_investment_data(association_name: Optional[str] = None) -> pd.DataFrame:
    """
    Get investment data for all associations or a specific association.
    
    Args:
        association_name: Optional association name to filter results
        
    Returns:
        DataFrame with investment data
    """
    query = f"""
    SELECT {TestDataColumns.BRANCH_NAME},
           {TestDataColumns.ASSOCIATION_NAME},
           {TestDataColumns.HOLDER},
           {TestDataColumns.INVESTMENT_TYPE},
           {TestDataColumns.INVESTMENT_TERM},
           {TestDataColumns.INVESTMENT_RATE},
           {TestDataColumns.CURRENT_BALANCE}
    FROM {TABLE_TEST_DATA}
    """
    
    params = []
    if association_name:
        query += f" WHERE {TestDataColumns.ASSOCIATION_NAME} = ?"
        params.append(association_name)
    
    query += f" ORDER BY {TestDataColumns.ASSOCIATION_NAME}"
    
    return execute_sql_query(query, params)

# Constants
SHEET_BANK_RANKING = 'Bank Ranking'
SHEET_BANK_RATES = 'Bank Rates'
SHEET_FILTER = 'Filter'
DEFAULT_MAX_BANK_ALLOCATION = 250000

# Term duration mappings (in months)
SHORT_TERM_MIN, SHORT_TERM_MAX = 1, 3
MID_TERM_MIN, MID_TERM_MAX = 4, 6
LONG_TERM_MIN, LONG_TERM_MAX = 7, 12

def load_data_from_database() -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Load and process data from the SQLite database.
        
    Returns:
        Tuple containing three DataFrames:
        - bank_ranking_df: DataFrame with bank rankings from test_data
        - bank_rates_df: DataFrame with CD rates for different terms
        - constraints_df: DataFrame with optimization constraints
    """
    try:
        # Get bank ranking data from test_data
        bank_ranking_query = f"""
        SELECT DISTINCT 
            {CDRatesColumns.BANK_NAME} as "Bank Name",
            COUNT(*) as "Priority"
        FROM {TABLE_CD_RATES}
        WHERE {CDRatesColumns.CD_RATE} IS NOT NULL
        AND {CDRatesColumns.BANK_NAME} IS NOT NULL
        AND {CDRatesColumns.BANK_NAME} != ''
        GROUP BY {CDRatesColumns.BANK_NAME}
        ORDER BY "Priority" DESC
        """
        bank_ranking_df = execute_sql_query(bank_ranking_query)
        
        # Get bank rates data
        bank_rates_query = f"""
        SELECT 
            {CDRatesColumns.BANK_NAME} as "Bank Name",
            {CDRatesColumns.CD_TERM} as "CD Term",
            {CDRatesColumns.CD_RATE} as "CD Rate"
        FROM {TABLE_CD_RATES}
        WHERE {CDRatesColumns.CD_RATE} IS NOT NULL
        AND {CDRatesColumns.BANK_NAME} IS NOT NULL
        AND {CDRatesColumns.BANK_NAME} != ''
        """
        bank_rates_df = execute_sql_query(bank_rates_query)
        
        # Clean and standardize CD terms
        bank_rates_df["CD Term"] = bank_rates_df["CD Term"].str.lower().str.strip()
        bank_rates_df["CD Term Num"] = bank_rates_df["CD Term"].str.extract(r'(\d+)').astype(float)

        # Remove any invalid terms
        bank_rates_df = bank_rates_df[bank_rates_df["CD Term Num"].notna()]
        
        # Create constraints DataFrame with default values
        constraints_data = {
            "Filter Name": ["Bank", "Term", "Rate"],
            "Min Value": [0, 1, 0],
            "Max Value": [250000, 12, 10]
        }
        constraints_df = pd.DataFrame(constraints_data)
        constraints_df.set_index("Filter Name", inplace=True)

        return bank_ranking_df, bank_rates_df, constraints_df
        
    except Exception as e:
        logger.error(f"Error loading data from database: {str(e)}")
        raise

def optimize_fund_allocation(
    bank_ranking_df: pd.DataFrame, 
    bank_rates_df: pd.DataFrame, 
    constraints: dict, 
    branch_name: str,
    association_name: str,
    custom_allocation: Optional[float] = None,  # Add parameter for custom allocation
    time_limit_seconds: int = 30
) -> Tuple[pd.DataFrame, float, float, float]:
    """
    Optimize fund allocation across banks and terms using a composite score based on:
    1. Category weights (sum to 100%) for each constraint category
    2. Individual constraint values (0-10) within each category
    3. Weighting factors that sum to 1.0 for interest vs ECR
    4. Liquidity reserve percentage (0-100%)
    
    Args:
        bank_ranking_df: DataFrame with bank ranking information
        bank_rates_df: DataFrame with rates information
        constraints: Dictionary with optimization constraints from database
        branch_name: Name of the branch to optimize for
        association_name: Name of the association to optimize for
        custom_allocation: Optional custom allocation amount to use instead of total funds
        time_limit_seconds: Time limit for the solver in seconds
        
    Returns:
        Tuple containing:
        - DataFrame with optimized allocation results
        - Float representing total funds available
    """
    try:
        # Get ECR rates
        ecr_rates = get_ecr_rates()
        
        solver = pywraplp.Solver.CreateSolver('GLOP')
        if not solver:
            raise Exception("GLOP solver is not available.")
            
        solver.set_time_limit(30000)  # 30 seconds in milliseconds
        
        # Get default max bank allocation
        max_bank_allocation = DEFAULT_MAX_BANK_ALLOCATION
        
        # Get bank relationships and filter rates
        branch_relationships = get_branch_relationships(branch_name)
        if branch_relationships.empty:
            raise ValueError(f"No relationships found for branch: {branch_name}")
            
        # Get list of banks with relationships (value = 1)
        related_banks = []
        # Map column names to actual bank names in the database
        bank_name_mapping = {
            'alliance_assoc_bank': 'Alliance Assoc. Bank',
            'banco_popular': 'Banco Popular',
            'bank_united': 'Bank United',
            'city_national': 'City National',
            'enterprise_bank_trust': 'Enterprise Bank & Trust',
            'first_citizens_bank': 'First Citizens Bank',
            'harmony_bank': 'Harmony Bank',
            'pacific_premier_bank': 'Pacific Premier Bank',
            'pacific_western': 'Pacific Western',
            'southstate': 'SouthState',
            'sunwest_bank': 'SunWest Bank',
            'capital_one': 'Capital One'
        }
        
        for bank_col in BranchRelationshipsColumns.__dict__.values():
            if isinstance(bank_col, str) and bank_col != BranchRelationshipsColumns.BRANCH_NAME:
                try:
                    if branch_relationships[bank_col].iloc[0] == 1:
                        # Use the mapping to get the correct bank name
                        bank_name = bank_name_mapping.get(bank_col, bank_col.replace('_', ' ').title())
                        related_banks.append(bank_name)
                        logger.info(f"Added related bank: {bank_name} (from column: {bank_col})")
                except KeyError:
                    continue
        
        if not related_banks:
            raise ValueError(f"No bank relationships found for branch: {branch_name}")
        
        # Filter bank rates to only include related banks
        filtered_rates = bank_rates_df[bank_rates_df["Bank Name"].isin(related_banks)].copy()
        
        if filtered_rates.empty:
            raise ValueError(f"No rates found for related banks of branch: {branch_name}")
        
        # 1. Get category weights directly from database (should sum to 100%)
        # Note: Liquidity constraints are handled separately and don't affect scoring
        category_weights = {
            'product': 0.0,
            'time': 0.0,
            'weighting': 0.0,
            'bank': 0.0
        }
        
        # Get category weights directly (these should already sum to 100%)
        for category in category_weights:
            if category in constraints:
                category_constraints = constraints[category]
                # Get the weight from the first constraint in the category
                if category_constraints:
                    first_constraint = next(iter(category_constraints.values()))
                    category_weights[category] = first_constraint['weight']
        
        # Verify total equals 100% (1.0)
        total_weight = sum(category_weights.values())
        if abs(total_weight - 1.0) > 0.01:
            logger.warning(f"Category weights sum to {total_weight*100:.1f}%, not 100%. Normalizing...")
            # Normalize to sum to 1.0 if they don't already
            if total_weight > 0:
                for category in category_weights:
                    category_weights[category] = category_weights[category] / total_weight
        
        logger.info(f"Category weights: {category_weights}")
        logger.info(f"Total category weight: {sum(category_weights.values())*100:.1f}%")
        logger.info("Note: Liquidity constraints are handled separately and don't affect scoring")
        
        # 2. Get bank weights (0-1 scale, already normalized)
        bank_weights = {}
        if 'bank' in constraints:
            bank_constraints = constraints['bank']
            for bank_name, bank_data in bank_constraints.items():
                if bank_data['enabled']:
                    # Values are already 0-1 scale
                    bank_weights[bank_name] = bank_data['value']
                    logger.info(f"Bank weight for {bank_name}: {bank_weights[bank_name]}")
        
        # 3. Get time weights (0-1 scale, already normalized)
        filtered_rates['time_weight'] = 0.0  # Default weight - start with 0 for disabled constraints
        if 'time' in constraints:
            time_constraints = constraints['time']
            time_mapping = {
                'Short Term (1-3 months)': (1, 3),
                'Mid Term (4-6 months)': (4, 6),
                'Long Term (7-12 months)': (7, 12)
            }
            
            for time_name, time_data in time_constraints.items():
                if time_data['enabled'] and time_name in time_mapping:
                    time_range = time_mapping[time_name]
                    # Values are already 0-1 scale
                    weight = time_data['value']
                    term_mask = (filtered_rates["CD Term Num"] >= time_range[0]) & (filtered_rates["CD Term Num"] <= time_range[1])
                    filtered_rates.loc[term_mask, 'time_weight'] = weight
                    logger.info(f"Time weight for {time_name}: {weight}")
                elif not time_data['enabled'] and time_name in time_mapping:
                    # Explicitly set weight to 0 for disabled constraints
                    time_range = time_mapping[time_name]
                    term_mask = (filtered_rates["CD Term Num"] >= time_range[0]) & (filtered_rates["CD Term Num"] <= time_range[1])
                    filtered_rates.loc[term_mask, 'time_weight'] = 0.0
                    logger.info(f"Time weight for {time_name} (DISABLED): 0.0")
        
        # 4. Get interest and ECR weights (sum to 1.0)
        interest_weight = 1.0
        ecr_weight = 0.0
        if 'weighting' in constraints:
            if 'Interest Rates' in constraints['weighting']:
                interest_data = constraints['weighting']['Interest Rates']
                if interest_data['enabled']:
                    interest_weight = interest_data['value']
                    
            if 'ECR Return' in constraints['weighting']:
                ecr_data = constraints['weighting']['ECR Return']
                if ecr_data['enabled']:
                    ecr_weight = ecr_data['value']
        
        # 5. Get liquidity reserve percentage (0-100%)
        liquidity_reserve = 0.3  # Default 30%
        if 'liquidity' in constraints:
            liquidity_constraints = constraints['liquidity']
            if liquidity_constraints:
                # Get the first liquidity constraint (should be "Liquidity Reserve")
                liquidity_name = next(iter(liquidity_constraints.keys()))
                liquidity_data = liquidity_constraints[liquidity_name]
                if liquidity_data['enabled']:
                    # Value is already stored as decimal (0.3 for 30%)
                    liquidity_reserve = liquidity_data['value']
                    logger.info(f"Using liquidity reserve: {liquidity_reserve*100:.1f}%")
                else:
                    logger.info("Liquidity constraint is disabled, using default 30%")
            else:
                logger.info("No liquidity constraints found, using default 30%")
        else:
            logger.info("Liquidity category not found in constraints, using default 30%")
        
        # Get total funds - either from database or use custom allocation
        if custom_allocation is not None:
            available_funds = custom_allocation
            total_funds = custom_allocation / (1 - liquidity_reserve)  # Calculate implied total funds
            logger.info(f"Using custom allocation amount: ${custom_allocation:,.2f}")
        else:
            # Get total funds from database (existing code)
            query = f"""
            SELECT SUM(current_balance) as total_balance
            FROM {TABLE_TEST_DATA}
            WHERE {TestDataColumns.ASSOCIATION_NAME} = '{association_name}'
            """
            balance_df = execute_sql_query(query)
            total_funds = float(balance_df['total_balance'].iloc[0] or 0) if not balance_df.empty else 0
            
            if total_funds <= 0:
                logger.warning(f"No funds available for association {association_name}")
                return pd.DataFrame(), total_funds

            available_funds = total_funds * (1 - liquidity_reserve)
        
        logger.info(f"Total funds: ${total_funds:,.2f}")
        logger.info(f"Liquidity reserve ({liquidity_reserve*100}%): ${total_funds * liquidity_reserve:,.2f}")
        logger.info(f"Available for allocation: ${available_funds:,.2f}")

        # Create allocation variables
        allocation = {}
        for bank in filtered_rates['Bank Name'].unique():
            bank_subset = filtered_rates[filtered_rates["Bank Name"] == bank]
            for _, row in bank_subset.iterrows():
                term = row["CD Term"]
                allocation[(bank, term)] = solver.IntVar(0, solver.infinity(), f'alloc_{bank}_{term}')
        
        # Add constraints
        solver.Add(sum(allocation.values()) <= available_funds)
        
        for bank in filtered_rates['Bank Name'].unique():
            vars_to_sum = [var for (b, _), var in allocation.items() if b == bank]
            if vars_to_sum:
                solver.Add(sum(vars_to_sum) <= max_bank_allocation)
        
        # Set objective function using composite score
        objective = solver.Objective()
        
        for (bank, term), var in allocation.items():
            rate_row = filtered_rates[(filtered_rates["Bank Name"] == bank) & 
                                    (filtered_rates["CD Term"] == term)]
            
            if not rate_row.empty:
                cd_rate = rate_row["CD Rate"].values[0] / 100
                bank_weight = bank_weights.get(bank, 1.0)
                time_weight = rate_row['time_weight'].values[0]
                
                # Calculate composite score
                composite_score = cd_rate * interest_weight
                
                # Add ECR component if enabled
                if ecr_weight > 0:
                    product_type = get_product_type(term)
                    ecr_rate = ecr_rates.get((bank, product_type), 0)
                    composite_score += ecr_rate * ecr_weight
                
                # Apply individual constraint weights
                # Bank weight (0-1, normalized from 0-10)
                bank_weight = bank_weights.get(bank, 1.0)
                
                # Time weight (0-1, normalized from 0-10)
                time_weight = rate_row['time_weight'].values[0]
                
                # Product weight for CDs (0-1 scale, already normalized)
                product_weight = 1.0  # Default weight
                if 'product' in constraints:
                    product_constraints = constraints['product']
                    if 'CD' in product_constraints and product_constraints['CD']['enabled']:
                        product_weight = product_constraints['CD']['value']
                
                # Apply category weights and individual constraint weights
                # This implements the 100% total system:
                # - Category weights determine overall influence (e.g., Product = 36%)
                # - Individual values determine contribution within category (e.g., CD = 100% of Product's 36% = 36% total)
                
                # Apply weighting category weight (interest vs ECR balance)
                if category_weights['weighting'] > 0:
                    composite_score *= category_weights['weighting']
                
                # Apply bank category weight and individual bank weight
                if category_weights['bank'] > 0:
                    composite_score *= (bank_weight * category_weights['bank'])
                else:
                    # If bank category has no weight, still apply individual bank preferences
                    composite_score *= bank_weight
                
                # Apply time category weight and individual time weight
                if category_weights['time'] > 0:
                    composite_score *= (time_weight * category_weights['time'])
                else:
                    # If time category has no weight, still apply individual time preferences
                    composite_score *= time_weight
                
                # Apply product category weight and individual product weight
                if category_weights['product'] > 0:
                    composite_score *= (product_weight * category_weights['product'])
                else:
                    # If product category has no weight, still apply individual product preferences
                    composite_score *= product_weight
                
                # Log the components of the score for debugging
                logger.info(f"Score components for {bank} {term}:")
                logger.info(f"  CD Rate: {cd_rate:.4f}, Interest Weight: {interest_weight:.3f}")
                logger.info(f"  Product: {product_weight:.3f} × {category_weights['product']:.3f} = {product_weight * category_weights['product']:.3f} ({product_weight * category_weights['product']*100:.1f}%)")
                logger.info(f"  Bank: {bank_weight:.3f} × {category_weights['bank']:.3f} = {bank_weight * category_weights['bank']:.3f} ({bank_weight * category_weights['bank']*100:.1f}%)")
                logger.info(f"  Time: {time_weight:.3f} × {category_weights['time']:.3f} = {time_weight * category_weights['time']:.3f} ({time_weight * category_weights['time']*100:.1f}%)")
                logger.info(f"  Weighting Category: {category_weights['weighting']:.3f} ({category_weights['weighting']*100:.1f}%)")
                logger.info(f"  Note: Liquidity constraints don't affect scoring, only fund availability")
                logger.info(f"  Final Composite Score: {composite_score:.6f}")
                
                objective.SetCoefficient(var, composite_score)
        
        objective.SetMaximization()
        
        # Solve and process results
        status = solver.Solve()
        
        if status != pywraplp.Solver.OPTIMAL:
            logger.warning(f"No optimal solution found. Solver status: {status}")
            return pd.DataFrame(), total_funds
        
        # Create results with actual rates for reporting
        results = []
        total_ecr = 0
        weighted_rate_sum = 0
        total_amount = 0
        
        for (bank, term), var in allocation.items():
            if var.solution_value() > 0:
                allocated_amount = int(var.solution_value())
                cd_rate = filtered_rates[(filtered_rates["Bank Name"] == bank) & 
                                      (filtered_rates["CD Term"] == term)]["CD Rate"].values[0]
                term_months = float(term.split()[0])
                expected_return = (allocated_amount * cd_rate * term_months) / (100 * 12)
                
                # Calculate ECR benefit
                product_type = get_product_type(term)
                ecr_rate = ecr_rates.get((bank, product_type), 0)
                estimated_ecr = allocated_amount * ecr_rate / 1200
                total_ecr += estimated_ecr
                
                # Calculate weighted rate (amount * rate)
                weighted_rate_sum += allocated_amount * cd_rate
                total_amount += allocated_amount
                
                results.append([
                    bank,
                    term,
                    allocated_amount,
                    cd_rate,
                    expected_return,
                    ecr_rate,
                    estimated_ecr
                ])
                  
        if not results:
            logger.warning("Optimization completed but no funds were allocated")
            logger.warning("Category weights (excluding liquidity):")
            for category, weight in category_weights.items():
                logger.warning(f"  {category}: {weight}")
            logger.warning("Bank weights:")
            for bank, weight in bank_weights.items():
                logger.warning(f"  {bank}: {weight}")
            logger.warning("Time weights:")
            logger.warning(filtered_rates[['CD Term', 'time_weight']].to_string())
            logger.warning("Note: Liquidity constraints only affect fund availability, not scoring")
            return pd.DataFrame(), total_funds, 0, 0
            
        columns = ["Bank Name", "CD Term", "Allocated Amount", "CD Rate", "Expected Return", "ECR Rate", "Estimated ECR Monthly"]
        df = pd.DataFrame(results, columns=columns)
        
        summary_row = pd.DataFrame([[
            "TOTAL",
            "",
            df["Allocated Amount"].sum(),
            None,
            df["Expected Return"].sum(),
            None,
            total_ecr
        ]], columns=columns)
        
        final_df = pd.concat([df, summary_row], ignore_index=True)
        
        return final_df, total_funds, weighted_rate_sum, total_amount
        
    except Exception as e:
        logger.error(f"Optimization error: {str(e)}")
        raise

def get_ecr_rates() -> Dict[Tuple[str, str], float]:
    """
    Get ECR rates from the database for all bank-product combinations.
    
    Returns:
        Dictionary mapping (bank_name, bank_code) tuples to ECR rates
    """
    try:
        query = """
        SELECT bank_name, bank_code, ecr_rate
        FROM ecr_rates
        WHERE ecr_rate IS NOT NULL
        AND ecr_rate > 0
        """
        df = execute_sql_query(query)
        
        # Create mapping of (bank_name, bank_code) to ecr_rate
        ecr_rates = {}
        for _, row in df.iterrows():
            ecr_rates[(row['bank_name'], row['bank_code'])] = float(row['ecr_rate'])
            
        return ecr_rates
        
    except Exception as e:
        logger.error(f"Error getting ECR rates: {str(e)}")
        return {}

def get_product_type(term: str) -> str:
    """
    Determine product type based on CD Term string.
    
    Args:
        term: CD Term string
        
    Returns:
        Product type code ('CD', 'CDARS', or 'MM')
    """
    if 'CDARS' in term:
        return 'CDARS'
    elif 'MM' in term:
        return 'MM'
    else:
        return 'CD'

def get_current_portfolio(association_name: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Get current portfolio data for an association from the database.
    
    Args:
        association_name: Name of the association
        
    Returns:
        Tuple containing:
        - List of current investments
        - Summary of current portfolio
    """
    try:
        # Get ECR rates
        ecr_rates = get_ecr_rates()
        logger.info(f"Fetching current portfolio for association: {association_name}")
        
        query = f"""
        SELECT 
            holder as bank,
            investment_type as product_type,
            current_balance as amount,
            investment_rate as rate,
            (current_balance * investment_rate / 1200) as monthly_interest
        FROM {TABLE_TEST_DATA}
        WHERE association_name = '{association_name}'
        AND current_balance > 0
        """
        
        logger.debug(f"Executing query: {query}")
        df = execute_sql_query(query)
        logger.info(f"Retrieved {len(df)} records for current portfolio.")
        
        if df.empty:
            logger.warning("No current portfolio data found.")
            return [], {'total_balance': 0, 'monthly_interest': 0, 'monthly_ecr': 0}
        
        current_portfolio = []
        total_ecr = 0
        
        for _, row in df.iterrows():
            # Determine product type and get ECR rate
            product_type = get_product_type(row['product_type'])
            ecr_rate = ecr_rates.get((row['bank'], product_type), 0)
            monthly_ecr = float(row['amount']) * ecr_rate / 1200
            total_ecr += monthly_ecr
            
            current_portfolio.append({
                'bank': row['bank'],
                'product_type': row['product_type'],
                'amount': float(row['amount']),
                'rate': float(row['rate']),
                'monthly_interest': float(row['monthly_interest']),
                'ecr_rate': ecr_rate,
                'monthly_ecr': monthly_ecr
            })
            
        summary = {
            'total_balance': sum(inv['amount'] for inv in current_portfolio),
            'monthly_interest': sum(inv['monthly_interest'] for inv in current_portfolio),
            'monthly_ecr': total_ecr
        }
        
        return current_portfolio, summary
        
    except Exception as e:
        logger.error(f"Error getting current portfolio: {str(e)}")
        return [], {'total_balance': 0, 'monthly_interest': 0, 'monthly_ecr': 0}

def run_optimization(params: Dict[str, Any]) -> Dict[str, Any]:
    """
    Run the optimization pipeline and return formatted results.
    
    Args:
        params: Dictionary containing optimization parameters including:
            - branch_name: Name of the branch
            - association_name: Name of the association
            - allocation_amount: Optional custom allocation amount
        
    Returns:
        Dictionary with optimization results and summary information
    """
    try:
        # Load data from database
        bank_ranking_df, bank_rates_df, _ = load_data_from_database()
        
        # Get constraints from database
        constraints = db_utils.get_constraints_for_optimizer()
        
        # Extract parameters with defaults
        branch_name = params.get('branch_name')
        association_name = params.get('association_name')
        allocation_amount = params.get('allocation_amount')  # Get custom allocation amount if provided
        
        if not branch_name or not association_name:
            return {
                'success': False,
                'message': 'Both branch name and association name are required for optimization.',
                'results': None
            }
            
        # Get current portfolio data
        current_portfolio, current_summary = get_current_portfolio(association_name)
        
        # Convert allocation_amount to float if provided
        if allocation_amount is not None:
            try:
                allocation_amount = float(allocation_amount)
                if allocation_amount <= 0:
                    return {
                        'success': False,
                        'message': 'Allocation amount must be greater than 0.',
                        'results': None
                    }
            except ValueError:
                return {
                    'success': False,
                    'message': 'Invalid allocation amount format.',
                    'results': None
                }
        
        # Run optimization with optional custom allocation amount
        result_df, total_funds, weighted_rate_sum, total_amount = optimize_fund_allocation(
            bank_ranking_df,
            bank_rates_df,
            constraints,
            branch_name,
            association_name,
            custom_allocation=allocation_amount
        )
        
        if result_df.empty:
            return {
                'success': False,
                'message': 'No allocation solution found with the given criteria.',
                'results': None
            }
            
        # Convert results to dictionary format
        results = result_df.to_dict(orient='records')
        
        # Extract summary row and make it separate
        summary = results[-1]
        results = results[:-1]
        
        return {
            'success': True,
            'message': 'Optimization completed successfully',
            'results': results,
            'current_portfolio': current_portfolio,
            'current_portfolio_summary': current_summary,
            'summary': {
                'total_allocated': summary['Allocated Amount'],
                'total_return': summary['Expected Return'],
                'weighted_avg_rate': (weighted_rate_sum / total_amount) if total_amount > 0 else 0,
                'total_funds': total_funds,
                'optimized_ecr_monthly': summary['Estimated ECR Monthly'],
                'current_ecr_monthly': current_summary['monthly_ecr'],
                'ecr_gain': summary['Estimated ECR Monthly'] - current_summary['monthly_ecr']
            },
            'bank_count': len(set(r['Bank Name'] for r in results)),
            'term_count': len(set(r['CD Term'] for r in results))
        }
        
    except Exception as e:
        logger.exception("Error in optimization pipeline")
        return {
            'success': False,
            'message': str(e),
            'results': None
        }

def export_results_to_excel(results: Dict[str, Any]) -> bytes:
    """
    Export optimization results to Excel file with a single comprehensive sheet.
    
    Args:
        results: Optimization results dictionary
        
    Returns:
        Excel file as bytes
    """
    if not results['success'] or not results['results']:
        df = pd.DataFrame([{'Error': results['message']}])
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return output.getvalue()

    # Create Excel writer
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create header section with dynamic values
        monthly_interest_increase = results['summary']['total_return'] - results['current_portfolio_summary']['monthly_interest']
        header_data = [
            ['', ''],  # Empty row for spacing
            [f"Hi {results.get('client_name', 'Client')},", ''],
            ['', ''],  # Empty row for spacing
            ['Thank you for taking the time to meet with me in reviewing your investment goals. Based on our conversation, I have completed a detailed', ''],
            [f"recommendation for you, which I've estimated to provide you up to ${results['summary']['total_return']:,.2f} in interest income per month which is an increase of ${monthly_interest_increase:,.2f} versus", ''],
            ['your current portfolio.', ''],
            ['', ''],  # Empty row for spacing
            ['Your current portfolio consists of:', ''],
            ['Funds Residing at', 'Product Type', '$ Invested', 'Current Rate of Return', 'Estimated Monthly Interest'],
        ]
        
        # Add current portfolio data
        current_portfolio = pd.DataFrame(header_data)
        
        # Add the current investments from results
        if results['current_portfolio']:
            for investment in results['current_portfolio']:
                current_portfolio.loc[len(current_portfolio)] = [
                    investment['bank'],
                    investment['product_type'],
                    f"${investment['amount']:,.2f}",
                    f"{investment['rate']:.3f}%",
                    f"${investment['monthly_interest']:,.2f}"
                ]
        
        # Add total row for current portfolio
        current_portfolio.loc[len(current_portfolio)] = [
            'Total Balance:',
            '',
            f"${results['current_portfolio_summary']['total_balance']:,.2f}",
            '',
            f"${results['current_portfolio_summary']['monthly_interest']:,.2f}"
        ]
        
        # Add spacing and proposal section
        current_portfolio.loc[len(current_portfolio)] = ['', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['Our proposal is as follows:', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['Send Funds to', 'Product Type', '$ Invested', 'Current Rate of Return', 'Estimated Monthly Interest']
        
        # Add proposed allocations
        for allocation in results['results']:
            current_portfolio.loc[len(current_portfolio)] = [
                allocation['Bank Name'],
                allocation['CD Term'],
                f"${allocation['Allocated Amount']:,.2f}",
                f"{allocation['CD Rate']:.3f}%",
                f"${allocation['Expected Return']:,.2f}"
            ]
        
        # Add total row for proposal
        current_portfolio.loc[len(current_portfolio)] = [
            'Total Balance:',
            '',
            f"${results['summary']['total_allocated']:,.2f}",
            '',
            f"${results['summary']['total_return']:,.2f}"
        ]
        
        # Generate dynamic CDARS recommendation text
        cdars_recommendations = []
        cdars_allocations = [alloc for alloc in results['results'] if 'CDARS' in alloc['CD Term']]
        if cdars_allocations:
            by_term = {}
            for alloc in cdars_allocations:
                term = alloc['CD Term']
                amount = alloc['Allocated Amount']
                if term not in by_term:
                    by_term[term] = []
                by_term[term].append(amount)
            
            for term, amounts in by_term.items():
                amounts_text = ', '.join(f"${amount:,.0f}k" for amount in amounts)
                cdars_recommendations.append(f"{term} CDARS for amounts {amounts_text}")
        
        cdars_text = "We recommend opening with PWB " + " and ".join(cdars_recommendations) + "." if cdars_recommendations else ""
        
        # Add notes section with dynamic content
        current_date = datetime.now().strftime('%m/%d/%Y')
        current_portfolio.loc[len(current_portfolio)] = ['', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['Notes:', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = [f"The amounts referenced in the current proposal section are from C3 balance as of {current_date}. Per our call, here is what the", '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['above proposal reflects:', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['1. We\'re going request a new PWB Money Market Reserve Account be opened with a $0 balance, this account will be used to move Solar', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['reserve funds in the future, and keep these funds segregated on your balance sheet.', '', '', '', '']
        if cdars_text:
            current_portfolio.loc[len(current_portfolio)] = ['2. ' + cdars_text, '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['All funds are 100% protected with maximum returns. Keep in mind that anything above can be changed at your discretion or with further', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['information.', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['To move forward with this proposal as outlined, please HAVE YOUR COMMUNITY MANAGER reply to this email and copy', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['clientservices@associa.us with the following information to allow us to assist you in the management of the account opening process:', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['1. Board Meeting minutes approving and outlining the desired plan of action', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['2. Articles of Incorporation', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['3. Signed Management Agreement', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['4. EIN/Federal Tax ID Number', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['5. Provide how many homes/units are in the community', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['6. Provide how much the average dues are, and the frequency of collection', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['Please note, based on the operating processes required to manage this task for you, accounts may take up to 2-3 weeks to open. This', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['timeline takes into consideration:', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['1. The SLA\'s of our internal bank ops team to make the account opening requests + validation of provided documentation', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['2. The SLA\'s of the bank\'s ops team to open the requested accounts', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['3. Account activation in the systems for management', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['4. The transference of funds, which is the responsibility of the Board and/or the CAM', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['5. Notice of completion', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['', '', '', '', '']
        current_portfolio.loc[len(current_portfolio)] = ['It\'s been our pleasure to serve you and thank you for giving us this opportunity to support your community\'s treasury needs.', '', '', '', '']
        
        # Write to Excel
        current_portfolio.to_excel(writer, sheet_name='Proposal', index=False, header=False)
        
        # Get the worksheet
        ws = writer.sheets['Proposal']
        
        # Format the worksheet
        # Set column widths
        ws.column_dimensions['A'].width = 30  # Funds Residing at / Send Funds to
        ws.column_dimensions['B'].width = 20  # Product Type
        ws.column_dimensions['C'].width = 15  # $ Invested
        ws.column_dimensions['D'].width = 25  # Current Rate of Return
        ws.column_dimensions['E'].width = 25  # Estimated Monthly Interest
        
        # Define styles
        header_fill = openpyxl.styles.PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')  # Light lavender
        alt_row_fill = openpyxl.styles.PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')  # Light gray
        total_fill = openpyxl.styles.PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')  # Light lavender
        bold_font = openpyxl.styles.Font(bold=True)
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        
        # Format current portfolio table
        current_portfolio_start = 9  # Row number where current portfolio table starts
        current_portfolio_end = current_portfolio_start + len(results['current_portfolio']) + 1  # Add 1 for the header row
        
        # Format headers
        for row in ws.iter_rows(min_row=current_portfolio_start, max_row=current_portfolio_start):
            for cell in row:
                cell.font = bold_font
                cell.fill = header_fill
                cell.border = thin_border
        
        # Format data rows
        for row_idx in range(current_portfolio_start + 1, current_portfolio_end):
            for cell in ws[row_idx]:
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_row_fill
        
        # Format total row
        for cell in ws[current_portfolio_end]:
            cell.font = bold_font
            cell.fill = total_fill
            cell.border = thin_border
        
        # Format proposal table
        proposal_start = current_portfolio_end + 3
        proposal_end = proposal_start + len(results['results']) + 1
        
        # Format proposal header
        for row in ws.iter_rows(min_row=proposal_start, max_row=proposal_start):
            for cell in row:
                cell.font = bold_font
                cell.fill = header_fill
                cell.border = thin_border
        
        # Format proposal data rows
        for row_idx in range(proposal_start + 1, proposal_end):
            for cell in ws[row_idx]:
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_row_fill
        
        # Format proposal total row
        for cell in ws[proposal_end]:
            cell.font = bold_font
            cell.fill = total_fill
            cell.border = thin_border
        
        # Add right alignment for numbers and rates
        for row in ws.iter_rows(min_row=current_portfolio_start):
            if row[2].value and isinstance(row[2].value, str) and row[2].value.startswith('$'):
                row[2].alignment = openpyxl.styles.Alignment(horizontal='right')
            if row[3].value and isinstance(row[3].value, str) and '%' in row[3].value:
                row[3].alignment = openpyxl.styles.Alignment(horizontal='right')
            if row[4].value and isinstance(row[4].value, str) and row[4].value.startswith('$'):
                row[4].alignment = openpyxl.styles.Alignment(horizontal='right')

    output.seek(0)
    return output.getvalue()

def analyze_portfolio(portfolio_path: str) -> Dict[str, Any]:
    """
    Analyze portfolio data against current rates in the database to identify optimization opportunities.
    Analysis is performed at the overall level, comparing current state with best available rates.
    
    Args:
        portfolio_path: Path to the portfolio Excel file
        
    Returns:
        Dictionary containing analysis results including:
        - Current portfolio metrics
        - Underperforming accounts
        - Rate analysis by term
        - Charts data for visualization
    """
    try:
        # Read portfolio data
        portfolio_df = pd.read_excel(portfolio_path)
        
        # Define column mapping for standardization
        column_mapping = {
            'summary_account': 'account',
            'summary_description': 'description',
            'pf_account_no': 'account_number',
            'account_desc': 'account_description',
            'holder': 'holder',
            'investment_type': 'product_type',
            'bank_account': 'bank_account',
            'purchase_date': 'purchase_date',
            'investment_term': 'term',
            'maturity_date': 'maturity_date',
            'investment_rate': 'rate',
            'note': 'note',
            'as_of_date': 'as_of_date',
            'current_balance': 'balance',
            'association_report_name': 'association_name'
        }
        
        # Clean and standardize column names
        portfolio_df.columns = [col.strip().lower().replace(' ', '_') for col in portfolio_df.columns]
        
        # Ensure required columns exist
        required_columns = ['current_balance', 'investment_rate', 'holder', 'investment_type']
        missing_columns = [col for col in required_columns if col not in portfolio_df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns in Excel file: {', '.join(missing_columns)}")
        
        # Rename columns using the mapping
        portfolio_df = portfolio_df.rename(columns=column_mapping)
        
        # Convert data types
        portfolio_df['balance'] = pd.to_numeric(portfolio_df['balance'], errors='coerce')
        portfolio_df['rate'] = pd.to_numeric(portfolio_df['rate'], errors='coerce')
        portfolio_df['term'] = pd.to_numeric(portfolio_df['term'], errors='coerce')
        
        # Remove rows with invalid data
        portfolio_df = portfolio_df.dropna(subset=['balance', 'rate'])
        
        # Get current rates from database
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get best available rates by term
        rate_query = """
        SELECT 
            cd_term,
            MAX(cd_rate) as best_rate
        FROM cd_rates
        WHERE cd_rate IS NOT NULL
        AND cd_rate > 0
        GROUP BY cd_term
        """
        cursor.execute(rate_query)
        best_rates = {row['cd_term']: float(row['best_rate']) for row in cursor.fetchall()}
        
        # Get partner bank relationships
        bank_query = """
        SELECT DISTINCT bank_name
        FROM cd_rates
        WHERE cd_rate IS NOT NULL
        AND cd_rate > 0
        """
        cursor.execute(bank_query)
        partner_banks = {row['bank_name'] for row in cursor.fetchall()}
        
        # Calculate current metrics
        total_balance = portfolio_df['balance'].sum()
        weighted_avg_rate = (portfolio_df['balance'] * portfolio_df['rate']).sum() / total_balance if total_balance > 0 else 0
        monthly_interest = (total_balance * weighted_avg_rate) / (12 * 100)
        annual_interest = monthly_interest * 12
        
        # Calculate best possible metrics (using best available rates)
        best_available_rate = max(best_rates.values())
        best_monthly_interest = (total_balance * best_available_rate) / (12 * 100)
        best_annual_interest = best_monthly_interest * 12
        
        # Calculate non-partner bank metrics
        non_partner_mask = ~portfolio_df['holder'].isin(partner_banks)
        non_partner_funds = portfolio_df.loc[non_partner_mask, 'balance'].sum()
        non_partner_pct = (non_partner_funds / total_balance * 100) if total_balance > 0 else 0
        
        # Identify underperforming accounts
        underperforming_mask = portfolio_df['rate'] < best_available_rate
        underperforming_funds = portfolio_df.loc[underperforming_mask, 'balance'].sum()
        underperforming_pct = (underperforming_funds / total_balance * 100) if total_balance > 0 else 0
        
        # Calculate rate metrics
        average_rate = weighted_avg_rate
        rate_gap = best_available_rate - average_rate
        potential_monthly_increase = best_monthly_interest - monthly_interest
        
        # Prepare underperforming accounts list
        underperforming_accounts = []
        for _, row in portfolio_df[underperforming_mask].iterrows():
            annual_loss = (row['balance'] * (best_available_rate - row['rate'])) / 100
            maturity_date = row.get('maturity_date', '')
            if pd.isna(maturity_date):
                maturity_date = ''
            else:
                try:
                    maturity_date = pd.to_datetime(maturity_date).strftime('%Y-%m-%d')
                except:
                    maturity_date = str(maturity_date)
            
            underperforming_accounts.append({
                'name': row['account'],
                'description': row.get('description', ''),
                'holder': row['holder'],
                'product_type': row['product_type'],
                'current_rate': row['rate'],
                'best_rate': best_available_rate,
                'balance': float(row['balance']),
                'annual_loss': annual_loss,
                'maturity_date': maturity_date,
                'recommendation': f"Consider reallocating to a {best_available_rate}% CD"
            })
        
        # Prepare rate analysis by term
        term_ranges = {
            'Short Term (1-3 months)': (1, 3),
            'Mid Term (4-6 months)': (4, 6),
            'Long Term (7-12 months)': (7, 12)
        }
        
        rate_analysis = []
        for term_name, (min_term, max_term) in term_ranges.items():
            term_mask = portfolio_df['term'].between(min_term, max_term)
            term_data = portfolio_df[term_mask]
            
            if not term_data.empty:
                term_balance = term_data['balance'].sum()
                term_avg_rate = (term_data['balance'] * term_data['rate']).sum() / term_balance if term_balance > 0 else 0
                
                # Get best rate for this term range
                term_best_rate = max((rate for term, rate in best_rates.items() 
                                    if min_term <= int(term.split()[0]) <= max_term), default=0)
                
                term_underperforming = term_data[term_data['rate'] < term_best_rate]['balance'].sum()
                potential_increase = (term_underperforming * (term_best_rate - term_avg_rate)) / 100
                
                rate_analysis.append({
                    'name': term_name,
                    'current_avg_rate': term_avg_rate,
                    'best_rate': term_best_rate,
                    'current_balance': term_balance,
                    'underperforming_funds': term_underperforming,
                    'potential_increase': potential_increase
                })
        
        # Prepare chart data
        rate_chart = {
            'labels': [term['name'] for term in rate_analysis],
            'current_rates': [term['current_avg_rate'] for term in rate_analysis],
            'best_rates': [term['best_rate'] for term in rate_analysis]
        }
        
        # Prepare maturity distribution data
        try:
            portfolio_df['maturity_date'] = pd.to_datetime(portfolio_df['maturity_date'], errors='coerce')
            maturity_bins = pd.date_range(start=pd.Timestamp.now(), periods=13, freq='ME')
            maturity_dist = pd.cut(portfolio_df['maturity_date'], bins=maturity_bins)
            maturity_amounts = portfolio_df.groupby(maturity_dist)['balance'].sum()
            
            maturity_chart = {
                'labels': [d.strftime('%b %Y') for d in maturity_bins[:-1]],
                'amounts': maturity_amounts.tolist()
            }
        except Exception as e:
            logger.warning(f"Error processing maturity dates: {str(e)}")
            maturity_chart = {
                'labels': [],
                'amounts': []
            }
        
        # Prepare non-partner distribution data
        non_partner_df = portfolio_df[non_partner_mask].groupby('holder')['balance'].sum()
        
        non_partner_chart = {
            'labels': non_partner_df.index.tolist(),
            'balances': non_partner_df.values.tolist()
        }
        
        conn.close()
        
        return {
            'non_partner_funds': non_partner_funds,
            'non_partner_change': non_partner_pct,
            'non_partner_change_direction': 'down',  # Always down as we want to reduce non-partner funds
            'non_partner_change_display': non_partner_pct,
            'underperforming_funds': underperforming_funds,
            'underperforming_change': underperforming_pct,
            'underperforming_change_direction': 'down',  # Always down as we want to reduce underperforming
            'underperforming_change_display': underperforming_pct,
            'average_rate': average_rate,
            'rate_change': abs(rate_gap),
            'rate_change_direction': 'up',  # Always up as we're showing potential improvement
            'rate_change_display': abs(rate_gap),
            'potential_rate_increase': rate_gap,
            'current_monthly_interest': monthly_interest,
            'best_monthly_interest': best_monthly_interest,
            'monthly_interest_increase': potential_monthly_increase,
            'current_annual_interest': annual_interest,
            'best_annual_interest': best_annual_interest,
            'annual_interest_increase': best_annual_interest - annual_interest,
            'underperforming': underperforming_accounts,
            'rate_analysis': rate_analysis,
            'rate_chart': rate_chart,
            'maturity_chart': maturity_chart,
            'non_partner_chart': non_partner_chart,
            'total_balance': total_balance,
            'total_accounts': len(portfolio_df),
            'accounts_underperforming': len(underperforming_accounts),
            'accounts_in_non_partner': len(non_partner_df)
        }
        
    except Exception as e:
        logger.exception("Error analyzing portfolio")
        raise ValueError(f"Error analyzing portfolio: {str(e)}")

def export_analysis_report(analysis_results: Dict[str, Any]) -> bytes:
    """
    Generate a detailed Excel report from the analysis results.
    
    Args:
        analysis_results: Dictionary containing analysis results
    
    Returns:
        Excel file as bytes
    """
    try:
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Summary Sheet
            summary_data = {
                'Metric': ['Total Balance', 'Monthly Interest', 'Annual Interest', 'Average Rate'],
                'Current': [
                    f"${analysis_results['total_balance']:,.2f}",
                    f"${analysis_results['current_monthly_interest']:,.2f}",
                    f"${analysis_results['current_annual_interest']:,.2f}",
                    f"{analysis_results['average_rate']:.3f}%"
                ],
                'Best Available': [
                    f"${analysis_results['total_balance']:,.2f}",
                    f"${analysis_results['best_monthly_interest']:,.2f}",
                    f"${analysis_results['best_annual_interest']:,.2f}",
                    f"{analysis_results['average_rate'] + analysis_results['potential_rate_increase']:.3f}%"
                ],
                'Potential Improvement': [
                    '-',
                    f"+${analysis_results['monthly_interest_increase']:,.2f}",
                    f"+${analysis_results['annual_interest_increase']:,.2f}",
                    f"+{analysis_results['potential_rate_increase']:.3f}%"
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            # Portfolio Overview Sheet
            overview_data = {
                'Metric': [
                    'Total Portfolio Balance',
                    'Total Number of Accounts',
                    'Non-Partner Bank Funds',
                    'Accounts in Non-Partner Banks',
                    'Underperforming Funds',
                    'Underperforming Accounts',
                    'Current Average Rate',
                    'Best Available Rate',
                    'Current Monthly Interest',
                    'Potential Monthly Interest',
                    'Monthly Interest Increase'
                ],
                'Value': [
                    f"${analysis_results['total_balance']:,.2f}",
                    analysis_results['total_accounts'],
                    f"${analysis_results['non_partner_funds']:,.2f} ({analysis_results['non_partner_change']:.1f}%)",
                    analysis_results['accounts_in_non_partner'],
                    f"${analysis_results['underperforming_funds']:,.2f} ({analysis_results['underperforming_change']:.1f}%)",
                    analysis_results['accounts_underperforming'],
                    f"{analysis_results['average_rate']:.3f}%",
                    f"{analysis_results['average_rate'] + analysis_results['potential_rate_increase']:.3f}%",
                    f"${analysis_results['current_monthly_interest']:,.2f}",
                    f"${analysis_results['best_monthly_interest']:,.2f}",
                    f"${analysis_results['monthly_interest_increase']:,.2f}"
                ]
            }
            pd.DataFrame(overview_data).to_excel(writer, sheet_name='Portfolio Overview', index=False)
            
            # Underperforming Accounts Sheet
            if analysis_results['underperforming']:
                underperforming_df = pd.DataFrame(analysis_results['underperforming'])
                underperforming_df = underperforming_df[[
                    'name', 'holder', 'product_type', 'current_rate', 'best_rate',
                    'balance', 'annual_loss', 'maturity_date', 'recommendation'
                ]]
                underperforming_df.columns = [
                    'Account', 'Bank/Holder', 'Product Type', 'Current Rate', 'Best Rate',
                    'Balance', 'Annual Loss', 'Maturity Date', 'Recommendation'
                ]
                underperforming_df.to_excel(writer, sheet_name='Underperforming Accounts', index=False)
            
            # Rate Analysis Sheet
            if analysis_results['rate_analysis']:
                rate_analysis_df = pd.DataFrame(analysis_results['rate_analysis'])
                rate_analysis_df = rate_analysis_df[[
                    'name', 'current_avg_rate', 'best_rate', 'current_balance',
                    'underperforming_funds', 'potential_increase'
                ]]
                rate_analysis_df.columns = [
                    'Term', 'Current Avg Rate', 'Best Rate', 'Current Balance',
                    'Underperforming Funds', 'Potential Annual Increase'
                ]
                rate_analysis_df.to_excel(writer, sheet_name='Rate Analysis', index=False)
            
            # Apply formatting
            workbook = writer.book
            
            # Format all sheets
            for sheet_name in writer.sheets:
                sheet = writer.sheets[sheet_name]
                
                # Set column widths
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                    if col in sheet.column_dimensions:
                        sheet.column_dimensions[col].width = 20
                
                # Add header style
                header_style = openpyxl.styles.NamedStyle(name=f'header_{sheet_name}')
                header_style.font = openpyxl.styles.Font(bold=True)
                header_style.fill = openpyxl.styles.PatternFill(
                    start_color='f1f8ff',
                    end_color='f1f8ff',
                    fill_type='solid'
                )
                
                # Apply header style to first row
                for cell in sheet[1]:
                    cell.style = header_style
                
                # Add number formatting
                for row in sheet.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, str):
                            if cell.value.startswith('$'):
                                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                            elif '%' in cell.value:
                                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
        
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        logger.exception("Error generating analysis report")
        raise Exception(f"Error generating analysis report: {str(e)}")

def get_db_connection() -> sqlite3.Connection:
    """
    Get a connection to the langston.db database.
    
    Returns:
        sqlite3.Connection: Database connection object
        
    Raises:
        Exception: If database connection fails
    """
    try:
        db_path = Path(__file__).parent.parent / "data" / "langston.db"
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row  # Enable row factory for named columns
        return conn
        
    except Exception as e:
        logger.error(f"Error connecting to database: {str(e)}")
        raise 